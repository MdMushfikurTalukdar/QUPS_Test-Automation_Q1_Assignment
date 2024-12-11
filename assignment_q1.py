import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
import time
import datetime

service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)

def read_keywords(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' does not exist.")
        return []
    sheet = wb[sheet_name]
    keywords = []
    
    for row in sheet.iter_rows(min_row=3, min_col=3, max_col=3, max_row=sheet.max_row):
        keyword = row[0].value
        if keyword:
            keywords.append(keyword)
        else:
            break
    return keywords

def today_day():
    today = datetime.datetime.today()
    return today.strftime('%A')

def write_results_to_excel(file_path, sheet_name, row, longest, shortest):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    sheet[f"D{row}"] = longest
    sheet[f"E{row}"] = shortest
    wb.save(file_path)



file_path = "Excel.xlsx" 

today_day = today_day()
#print(f"Today's day is: {today_day}")

wb = openpyxl.load_workbook(file_path)
sheets = wb.sheetnames
if today_day in sheets:
    sheet_name = today_day
else:
    sheet_name = "Saturday"

keywords = read_keywords(file_path, sheet_name)

if keywords:
    row_number = 3
    for keyword in keywords:
        driver.get("https://www.google.com")
        search_box = driver.find_element(By.NAME, "q")
        search_box.send_keys(keyword)
        search_box.send_keys(Keys.RETURN)
        time.sleep(3)

        search_results = driver.find_elements(By.CSS_SELECTOR, "h3")

        results = [result.text for result in search_results if result.text.strip() != ""]

        if results:
            longest_title = max(results, key=lambda result: len(result))
            shortest_title = min(results, key=lambda result: len(result))
            # print("Longest Search Result Title: ", longest_title)
            # print("Shortest Search Result Title: ", shortest_title)

            write_results_to_excel(file_path, sheet_name, row_number, longest_title, shortest_title)
            
            row_number += 1
        else:
            print(f"No valid search results found for '{keyword}'.")
else:
    print("No keywords found in the selected sheet.")

print("Excel sheed updated")
driver.quit()